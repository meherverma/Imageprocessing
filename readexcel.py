import openpyxl

path='C://Users//meher.verma//Desktop//read1.xlsx'

workbook=openpyxl.load_workbook(path)

#sheet_num=workbook.get_sheet_by_name('Sheet1')
sheet_num=workbook.active

rows=sheet_num.max_row
cols=sheet_num.max_column

print(rows)
print(cols)

for i in range(1,rows+1):
    for j in range(1, cols+1):
        print(sheet_num.cell(row=i, column=j).value, end="  ")
    print()

for i in range(15, 17):
    for j in range(1, cols + 1):
        sheet_num.cell(row=i, column=j).value="Meher"

workbook.save(path)



